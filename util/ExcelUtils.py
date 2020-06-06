import os

import openpyxl

from util import Logger


def readdata(filename, sheetname, rowno):
    log = Logger.getlogger()
    try:
        path = os.getcwd()
        projectroot = os.path.dirname(path)
        path = projectroot + "\\resources\\testdata\\"+filename
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]

        mydict = {}

        for colno in range(1, sheet.max_column + 1):
            mydict[sheet.cell(row=1, column=colno).value] = sheet.cell(row=rowno, column=colno).value

        log.info("Data is read successfully.")

        return mydict

    except Exception as e:
        log.error("Exception Occurred", exc_info=True)
    finally:
        workbook.close()
        log.info("Workbook closed succesfully")


def getdata(mydict, colname):
    return mydict[colname]


def writedata(filename, sheetname, rowno, mydict):
    log = Logger.getlogger()
    try:
        path = os.getcwd()
        projectroot = os.path.dirname(path)
        path = projectroot + "\\resources\\testdata\\" + filename
        workbook = openpyxl.load_workbook(path)
        sheet = workbook[sheetname]

        for colno in range(1, sheet.max_column + 1):
            key = str(sheet.cell(row=1, column=colno).value)
            if key in mydict:
                sheet.cell(row=rowno, column=colno).value = ""
                sheet.cell(row=rowno, column=colno).value = mydict[key]

        log.info("Data is write successfully.")

    except Exception as e:
        log.error("Exception Occurred", exc_info=True)
    finally:
        workbook.save(path)
        workbook.close()
        log.info("Workbook closed successfully.")

